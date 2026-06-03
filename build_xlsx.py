# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter
import datetime

OUT = "/Users/johannlemena/Desktop/Plan_30k_LED_Juin/Plan_30k_LED_Juin.xlsx"

# ---------- Palette ----------
NAVY   = "1F3A5F"   # bandeaux titres
BLUE   = "2E6FB7"   # accents
LIGHT  = "EAF1FB"   # fond doux
GREY   = "F2F2F2"
GREEN  = "2E7D32"
RED    = "C0392B"
AMBER  = "E08600"
WHITE  = "FFFFFF"
HEADERFILL = PatternFill("solid", fgColor=NAVY)
SUBFILL    = PatternFill("solid", fgColor=BLUE)
LIGHTFILL  = PatternFill("solid", fgColor=LIGHT)
GREYFILL   = PatternFill("solid", fgColor=GREY)
INPUTFILL  = PatternFill("solid", fgColor="FFF7D6")  # jaune doux = à saisir

F = "Arial"
def font(sz=10, b=False, color="000000", it=False): return Font(name=F, size=sz, bold=b, color=color, italic=it)
thin = Side(style="thin", color="BFBFBF")
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
LFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RGT = Alignment(horizontal="right", vertical="center")

wb = Workbook()

# ============================================================
# 1) PARAMETRES
# ============================================================
p = wb.active
p.title = "PARAMETRES"
p.sheet_view.showGridLines = False
p["A1"] = "PARAMETRES — Objectif & configuration"
p["A1"].font = font(14, True, WHITE); p["A1"].fill = HEADERFILL; p["A1"].alignment = LFT
p.merge_cells("A1:F1"); p.row_dimensions[1].height = 26

cfg = [
    ("Objectif LED (mois)", 30000, "#,##0"),
    ("Cadence cible (LED / jour)", 1500, "#,##0"),
    ("Date de debut de periode", datetime.date(2026,6,1), "dd/mm/yyyy"),
    ("Date de fin de periode", datetime.date(2026,6,30), "dd/mm/yyyy"),
    ("Aujourd'hui (auto)", "=TODAY()", "dd/mm/yyyy"),
    ("Jours ouvres restants (auto)", "=NETWORKDAYS(B7,B6)", "0"),
    ("Seuil de completude mini pour DEPOSER", 1.0, "0%"),
    ("Prime equipe si objectif atteint (EUR / pers)", 200, "#,##0"),
]
r = 3
for label, val, fmt in cfg:
    p.cell(r,1,label).font = font(10, True)
    c = p.cell(r,2,val); c.font = font(10, color="0000FF" if not str(val).startswith("=") else "000000")
    c.number_format = fmt; c.alignment = RGT; c.fill = INPUTFILL if not str(val).startswith("=") else GREYFILL
    c.border = BORD; p.cell(r,1).border = BORD
    r += 1

# listes pour menus deroulants
lists = {
    "D": ("Statuts", ["A controler","En controle","A corriger","Pret a deposer","Depose","Bloque"]),
    "E": ("Motifs de blocage", ["RAS","Attente pose client","Attente justif client","Attente Pappers","Attente devis resigne","Attente audit/synthese","Doute conformite"]),
    "F": ("Responsables", ["Johann","Coequipier 1","Coequipier 2"]),
}
for col,(title,vals) in lists.items():
    p[col+"11"] = title; p[col+"11"].font = font(10, True, WHITE); p[col+"11"].fill = SUBFILL; p[col+"11"].alignment = CTR; p[col+"11"].border = BORD
    for i,v in enumerate(vals):
        cc = p[col+str(12+i)]; cc.value = v; cc.font = font(10); cc.border = BORD
p.column_dimensions["A"].width = 38
for c in "BDEF": p.column_dimensions[c].width = 22
p.column_dimensions["C"].width = 4

# ============================================================
# 2) PIPELINE
# ============================================================
pl = wb.create_sheet("PIPELINE")
pl.sheet_view.showGridLines = False
pl["A1"] = "PIPELINE DOSSIERS — coller ici l'export Pixel/Betool (1 ligne = 1 dossier).  Donnees d'amorcage = exemples connus a remplacer."
pl["A1"].font = font(11, True, WHITE); pl["A1"].fill = HEADERFILL; pl["A1"].alignment = LFT
pl.merge_cells("A1:Q1"); pl.row_dimensions[1].height = 24

headers = ["ID","Entreprise","SIRET","Secteur","Nb chantiers","Nb LED","Prime CEE (EUR TTC)",
           "Surface (m2)","Date signature","% Pose reelle","Statut","Motif blocage","Responsable",
           "Date depot","Age (j)","Priorite","Action du jour"]
for j,h in enumerate(headers, start=1):
    c = pl.cell(2,j,h); c.font = font(9, True, WHITE); c.fill = SUBFILL; c.alignment = CTR; c.border = BORD
pl.row_dimensions[2].height = 32
pl.freeze_panes = "A3"

# Donnees d'amorcage (issues du .md et des fichiers Downloads)
# Donnees d'amorcage ANONYMISEES (exemples) — a remplacer par l'export CRM (jamais de PII reelle ici)
seed = [
    ("Exemple AGRI 1",None,"AGRI",3,60,3900,3206,datetime.date(2026,5,25),1.00,"A corriger","RAS","Johann"),
    ("Exemple AGRI 2",None,"AGRI",3,114,7410,4497,datetime.date(2026,5,20),1.00,"A controler","RAS","Coequipier 1"),
    ("Exemple BTP 1",None,"Construction",2,36,2340,1453,datetime.date(2026,5,15),0.70,"Bloque","Attente pose client","Coequipier 2"),
    ("Exemple AGRI 3",None,"AGRI",1,36,None,1850,datetime.date(2026,5,19),1.00,"Pret a deposer","RAS","Johann"),
    ("Exemple AGRI 4",None,"AGRI",2,36,2340,2235,datetime.date(2026,5,28),0.65,"Bloque","Attente pose client","Coequipier 1"),
]
FIRST = 3
LAST  = 302  # capacite 300 dossiers
for i,row in enumerate(seed):
    r = FIRST + i
    ent,siret,sect,nbch,led,prime,surf,dsig,pose,stat,motif,resp = row
    pl.cell(r,2,ent); pl.cell(r,3,siret); pl.cell(r,4,sect); pl.cell(r,5,nbch)
    pl.cell(r,6,led); pl.cell(r,7,prime); pl.cell(r,8,surf); pl.cell(r,9,dsig)
    pl.cell(r,10,pose); pl.cell(r,11,stat); pl.cell(r,12,motif); pl.cell(r,13,resp)

# formules colonnes calculees pour toutes les lignes (3..302)
for r in range(FIRST, LAST+1):
    pl.cell(r,1).value  = '=IF(B{r}="","",ROW()-2)'.format(r=r)                         # ID
    pl.cell(r,15).value = '=IF(I{r}="","",TODAY()-I{r})'.format(r=r)                     # Age
    # Priorite = base statut + completude + LED + age, tie-breaker unique
    pl.cell(r,16).value = (
        '=IF(B{r}="","",'
        'IF(K{r}="Pret a deposer",400,IF(K{r}="A corriger",300,IF(K{r}="En controle",250,'
        'IF(K{r}="A controler",200,IF(K{r}="Bloque",60,IF(K{r}="Depose",0,150))))))'
        '+ N(J{r})*60 + N(F{r})*0.5 + IFERROR(N(O{r}),0)*0.1 + ROW()/100000)'
    ).format(r=r)
    # Action du jour
    pl.cell(r,17).value = (
        '=IF(B{r}="","",'
        'IF(K{r}="Depose","OK Depose",'
        'IF(AND(J{r}<>"",J{r}<\'PARAMETRES\'!$B$9),"Relancer pose ("&TEXT(J{r},"0%")&")",'
        'IF(K{r}="Pret a deposer","=> DEPOSER aujourd hui",'
        'IF(K{r}="A corriger","Corriger + deposer",'
        'IF(K{r}="En controle","Finir le controle",'
        'IF(K{r}="A controler","Lancer le controle",'
        'IF(K{r}="Bloque","Debloquer: "&L{r},""))))))))'
    ).format(r=r)

# largeurs
widths = {"A":5,"B":26,"C":15,"D":12,"E":9,"F":8,"G":15,"H":11,"I":13,"J":10,
          "K":14,"L":21,"M":13,"N":12,"O":7,"P":9,"Q":22}
for col,w in widths.items(): pl.column_dimensions[col].width = w

# formats nombres / style data zone
for r in range(FIRST, LAST+1):
    pl.cell(r,6).number_format = "#,##0"
    pl.cell(r,7).number_format = '#,##0" EUR"'
    pl.cell(r,8).number_format = '#,##0" m2"'
    pl.cell(r,9).number_format = "dd/mm/yyyy"
    pl.cell(r,10).number_format = "0%"
    pl.cell(r,14).number_format = "dd/mm/yyyy"
    pl.cell(r,15).number_format = '0" j"'
    pl.cell(r,16).number_format = "0.0"
    pl.cell(r,10).fill = INPUTFILL  # % pose = saisie cle
    for j in range(1,18):
        cc = pl.cell(r,j); cc.font = font(9); cc.border = BORD
        if j in (1,5,6,15,16): cc.alignment = CTR
        elif j in (7,8): cc.alignment = RGT
        else: cc.alignment = LFT
    if (r-FIRST) % 2 == 1:
        for j in range(1,18):
            if pl.cell(r,j).fill.fgColor.rgb in (None,"00000000"):
                pl.cell(r,j).fill = GREYFILL

# menus deroulants
dv_stat = DataValidation(type="list", formula1="='PARAMETRES'!$D$12:$D$17", allow_blank=True)
dv_mot  = DataValidation(type="list", formula1="='PARAMETRES'!$E$12:$E$18", allow_blank=True)
dv_resp = DataValidation(type="list", formula1="='PARAMETRES'!$F$12:$F$14", allow_blank=True)
pl.add_data_validation(dv_stat); pl.add_data_validation(dv_mot); pl.add_data_validation(dv_resp)
dv_stat.add("K{}:K{}".format(FIRST,LAST))
dv_mot.add("L{}:L{}".format(FIRST,LAST))
dv_resp.add("M{}:M{}".format(FIRST,LAST))

# mise en forme conditionnelle : % pose < seuil -> rouge ; statut depose -> vert
pl.conditional_formatting.add("J{}:J{}".format(FIRST,LAST),
    CellIsRule(operator="lessThan", formula=["1"], fill=PatternFill("solid", fgColor="F8CBAD"), font=font(9, color=RED)))
pl.conditional_formatting.add("K{}:K{}".format(FIRST,LAST),
    FormulaRule(formula=['$K{}="Depose"'.format(FIRST)], fill=PatternFill("solid", fgColor="C6EFCE"), font=font(9, color=GREEN)))
pl.conditional_formatting.add("K{}:K{}".format(FIRST,LAST),
    FormulaRule(formula=['$K{}="Bloque"'.format(FIRST)], fill=PatternFill("solid", fgColor="FCE4D6"), font=font(9, color=AMBER)))

# ============================================================
# 3) DASHBOARD
# ============================================================
d = wb.create_sheet("DASHBOARD", 0)
d.sheet_view.showGridLines = False
PL = "PIPELINE!"
LEDr = PL+"$F$3:$F$302"; STATr = PL+"$K$3:$K$302"; MOTr = PL+"$L$3:$L$302"
ENTr = PL+"$B$3:$B$302"; PRIMr = PL+"$G$3:$G$302"; SURFr = PL+"$H$3:$H$302"
POSEr= PL+"$J$3:$J$302"; PRIOr = PL+"$P$3:$P$302"; ACTr = PL+"$Q$3:$Q$302"

def title_band(rng, text, size=14):
    first = rng.split(":")[0]
    d[first] = text; d[first].font = font(size, True, WHITE); d[first].fill = HEADERFILL; d[first].alignment = LFT
    d.merge_cells(rng)

title_band("A1:L1", "TABLEAU DE BORD — OBJECTIF 30 000 LED · JUIN 2026")
d.row_dimensions[1].height = 28
d["A2"] = "Cockpit quotidien — a actualiser chaque matin (saisir la pose dans PIPELINE + les LED deposees du jour ci-dessous)."
d["A2"].font = font(9, it=True, color="555555"); d.merge_cells("A2:L2")
d["A3"] = "Aujourd'hui :"; d["A3"].font = font(9, True); d["B3"] = "=TODAY()"; d["B3"].number_format="dddd dd/mm/yyyy"; d["B3"].font=font(9,True,color=BLUE)

# ---- bandeau KPI (rangs 5-6) ----
kpis = [
    ("Objectif mois", "='PARAMETRES'!B3", "#,##0"),
    ("Depose (cumul)", "=SUMIFS({L},{S},\"Depose\")".format(L=LEDr,S=STATr), "#,##0"),
    ("Reste a deposer", "=B6-A6 + 0", "#,##0"),  # placeholder, set below explicitly
    ("Jours ouvres restants", "='PARAMETRES'!B8", "0"),
    ("Rythme requis / j", None, "#,##0"),
    ("Cadence cible / j", "='PARAMETRES'!B4", "#,##0"),
    ("Pret a deposer (LED)", "=SUMIFS({L},{S},\"Pret a deposer\")".format(L=LEDr,S=STATr), "#,##0"),
    ("Completude moy. pond.", None, "0%"),
]
cols = ["A","B","C","D","E","F","G","H"]
for i,(lab,form,fmt) in enumerate(kpis):
    col = cols[i]
    lc = d[col+"5"]; lc.value = lab; lc.font = font(9, True, WHITE); lc.fill = SUBFILL; lc.alignment = CTR; lc.border = BORD
    vc = d[col+"6"]; vc.number_format = fmt; vc.font = font(15, True, color=NAVY); vc.alignment = CTR; vc.fill = LIGHTFILL; vc.border = BORD
d.row_dimensions[6].height = 30
d["A6"] = "='PARAMETRES'!B3"
d["B6"] = "=SUMIFS({L},{S},\"Depose\")".format(L=LEDr,S=STATr)
d["C6"] = "=MAX(0,A6-B6)"
d["D6"] = "='PARAMETRES'!B8"
d["E6"] = "=IFERROR(ROUNDUP(C6/D6,0),0)"
d["F6"] = "='PARAMETRES'!B4"
d["G6"] = "=SUMIFS({L},{S},\"Pret a deposer\")".format(L=LEDr,S=STATr)
d["H6"] = "=IFERROR(SUMPRODUCT(({P}<>\"\")*({P})*({L}))/SUMPRODUCT(({P}<>\"\")*({L})),\"\")".format(P=POSEr,L=LEDr)

# ---- FUNNEL par statut (A8:C16) ----
d["A8"]="FUNNEL — par statut"; d["A8"].font=font(11,True,WHITE); d["A8"].fill=SUBFILL; d.merge_cells("A8:C8"); d["A8"].alignment=LFT
for j,h in enumerate(["Statut","Nb dossiers","LED"]):
    c=d.cell(9,1+j,h); c.font=font(9,True); c.fill=GREYFILL; c.alignment=CTR; c.border=BORD
statuts=["A controler","En controle","A corriger","Pret a deposer","Depose","Bloque"]
for i,s in enumerate(statuts):
    r=10+i
    d.cell(r,1,s).font=font(9); d.cell(r,1).border=BORD
    d.cell(r,2,"=COUNTIFS({S},A{r})".format(S=STATr,r=r)).border=BORD
    d.cell(r,3,"=SUMIFS({L},{S},A{r})".format(L=LEDr,S=STATr,r=r)).border=BORD
    d.cell(r,2).alignment=CTR; d.cell(r,3).number_format="#,##0"; d.cell(r,3).alignment=RGT; d.cell(r,3).font=font(9)
d.cell(16,1,"TOTAL").font=font(9,True); d.cell(16,1).border=BORD
d.cell(16,2,"=SUM(B10:B15)").font=font(9,True); d.cell(16,2).alignment=CTR; d.cell(16,2).border=BORD
d.cell(16,3,"=SUM(C10:C15)").font=font(9,True); d.cell(16,3).number_format="#,##0"; d.cell(16,3).alignment=RGT; d.cell(16,3).border=BORD

# ---- BLOCAGES par motif (E8:G16) ----
d["E8"]="BLOCAGES — par motif"; d["E8"].font=font(11,True,WHITE); d["E8"].fill=SUBFILL; d.merge_cells("E8:G8"); d["E8"].alignment=LFT
for j,h in enumerate(["Motif","Nb","LED"]):
    c=d.cell(9,5+j,h); c.font=font(9,True); c.fill=GREYFILL; c.alignment=CTR; c.border=BORD
motifs=["Attente pose client","Attente justif client","Attente Pappers","Attente devis resigne","Attente audit/synthese","Doute conformite"]
for i,m in enumerate(motifs):
    r=10+i
    d.cell(r,5,m).font=font(9); d.cell(r,5).border=BORD
    d.cell(r,6,"=COUNTIFS({M},E{r})".format(M=MOTr,r=r)).border=BORD; d.cell(r,6).alignment=CTR
    d.cell(r,7,"=SUMIFS({L},{M},E{r})".format(L=LEDr,M=MOTr,r=r)).border=BORD; d.cell(r,7).number_format="#,##0"; d.cell(r,7).alignment=RGT; d.cell(r,7).font=font(9)
d.cell(16,5,"TOTAL bloque").font=font(9,True); d.cell(16,5).border=BORD
d.cell(16,6,"=SUM(F10:F15)").font=font(9,True); d.cell(16,6).alignment=CTR; d.cell(16,6).border=BORD
d.cell(16,7,"=SUM(G10:G15)").font=font(9,True); d.cell(16,7).number_format="#,##0"; d.cell(16,7).alignment=RGT; d.cell(16,7).border=BORD

# ---- PROFONDEUR (I8:J20) ----
d["I8"]="PROFONDEUR DU PIPELINE"; d["I8"].font=font(11,True,WHITE); d["I8"].fill=SUBFILL; d.merge_cells("I8:J8"); d["I8"].alignment=LFT
prof=[
    ("Nb dossiers","=COUNTA({E})".format(E=ENTr),"0"),
    ("Total LED (tout statut)","=SUM({L})".format(L=LEDr),"#,##0"),
    ("Total LED hors depose","=SUMIFS({L},{S},\"<>Depose\")".format(L=LEDr,S=STATr),"#,##0"),
    ("Moyenne LED / dossier","=IFERROR(AVERAGEIF({L},\">0\"),\"\")".format(L=LEDr),"#,##0.0"),
    ("Mediane LED","=IFERROR(MEDIAN({L}),\"\")".format(L=LEDr),"#,##0"),
    ("Min LED","=IFERROR(MIN({L}),\"\")".format(L=LEDr),"#,##0"),
    ("Max LED","=IFERROR(MAX({L}),\"\")".format(L=LEDr),"#,##0"),
    ("Moyenne prime (EUR)","=IFERROR(AVERAGEIF({G},\">0\"),\"\")".format(G=PRIMr),"#,##0"),
    ("Moyenne surface (m2)","=IFERROR(AVERAGEIF({H},\">0\"),\"\")".format(H=SURFr),"#,##0"),
    ("Completude moy. ponderee","=H6","0%"),
    ("Dossiers < seuil pose","=COUNTIFS({P},\"<\"&'PARAMETRES'!$B$9,{P},\"<>\")".format(P=POSEr),"0"),
]
for i,(lab,form,fmt) in enumerate(prof):
    r=9+i
    d.cell(r,9,lab).font=font(9); d.cell(r,9).border=BORD; d.cell(r,9).alignment=LFT
    c=d.cell(r,10,form); c.font=font(9,True,color=NAVY); c.number_format=fmt; c.alignment=RGT; c.border=BORD

# ---- SUIVI QUOTIDIEN burn-up (A19:E50) ----
d["A21"]="SUIVI QUOTIDIEN (burn-up) — saisir les LED deposees chaque jour (colonne jaune)"
d["A21"].font=font(11,True,WHITE); d["A21"].fill=SUBFILL; d.merge_cells("A21:E21"); d["A21"].alignment=LFT
for j,h in enumerate(["Date","LED deposees (saisie)","Cumul reel","Cumul cible","Ecart"]):
    c=d.cell(22,1+j,h); c.font=font(9,True); c.fill=GREYFILL; c.alignment=CTR; c.border=BORD
JR0=23
for i in range(30):
    r=JR0+i
    if i==0: d.cell(r,1,"='PARAMETRES'!B5")
    else: d.cell(r,1,"=A{}+1".format(r-1))
    d.cell(r,1).number_format="ddd dd/mm"; d.cell(r,1).font=font(9); d.cell(r,1).border=BORD; d.cell(r,1).alignment=CTR
    inp=d.cell(r,2); inp.fill=INPUTFILL; inp.border=BORD; inp.number_format="#,##0"; inp.alignment=CTR; inp.font=font(9,color="0000FF")
    d.cell(r,3,"=SUM($B${0}:B{1})".format(JR0,r)).number_format="#,##0"; d.cell(r,3).border=BORD; d.cell(r,3).alignment=RGT; d.cell(r,3).font=font(9)
    d.cell(r,4,"=MIN('PARAMETRES'!$B$3,'PARAMETRES'!$B$4*NETWORKDAYS('PARAMETRES'!$B$5,A{r}))".format(r=r)).number_format="#,##0"; d.cell(r,4).border=BORD; d.cell(r,4).alignment=RGT; d.cell(r,4).font=font(9)
    d.cell(r,5,"=C{r}-D{r}".format(r=r)).number_format="#,##0;[Red]-#,##0"; d.cell(r,5).border=BORD; d.cell(r,5).alignment=RGT; d.cell(r,5).font=font(9)
d.conditional_formatting.add("E{}:E{}".format(JR0,JR0+29),
    CellIsRule(operator="lessThan", formula=["0"], font=font(9,color=RED)))

# ---- TOP a traiter (G19:K34) avec colonne helper M ----
d["G21"]="TOP 12 — a traiter aujourd'hui (par priorite)"; d["G21"].font=font(11,True,WHITE); d["G21"].fill=SUBFILL; d.merge_cells("G21:K21"); d["G21"].alignment=LFT
for j,h in enumerate(["#","Entreprise","LED","Statut","Action"]):
    c=d.cell(22,7+j,h); c.font=font(9,True); c.fill=GREYFILL; c.alignment=CTR; c.border=BORD
for i in range(12):
    r=23+i; rang=i+1
    d.cell(r,7,rang).font=font(9,True); d.cell(r,7).alignment=CTR; d.cell(r,7).border=BORD
    d.cell(r,13,"=IFERROR(MATCH(LARGE({P},G{r}),{P},0),\"\")".format(P=PRIOr,r=r))  # helper M
    d.cell(r,8,"=IFERROR(INDEX({E},$M{r}),\"\")".format(E=ENTr,r=r)).border=BORD; d.cell(r,8).font=font(9); d.cell(r,8).alignment=LFT
    d.cell(r,9,"=IFERROR(INDEX({L},$M{r}),\"\")".format(L=LEDr,r=r)).border=BORD; d.cell(r,9).number_format="#,##0"; d.cell(r,9).alignment=CTR; d.cell(r,9).font=font(9)
    d.cell(r,10,"=IFERROR(INDEX({S},$M{r}),\"\")".format(S=STATr,r=r)).border=BORD; d.cell(r,10).font=font(9); d.cell(r,10).alignment=LFT
    d.cell(r,11,"=IFERROR(INDEX({A},$M{r}),\"\")".format(A=ACTr,r=r)).border=BORD; d.cell(r,11).font=font(9); d.cell(r,11).alignment=LFT
d.column_dimensions["M"].hidden = True

# largeurs dashboard
dwidths={"A":20,"B":15,"C":11,"D":12,"E":13,"F":13,"G":11,"H":12,"I":22,"J":12,"K":22,"L":4}
for col,w in dwidths.items(): d.column_dimensions[col].width=w

# graphique burn-up
chart=LineChart(); chart.title="Burn-up — Cumul reel vs cible"; chart.height=7.5; chart.width=15
chart.style=2
cats=Reference(d, min_col=1, min_row=JR0, max_row=JR0+29)
data=Reference(d, min_col=3, max_col=4, min_row=22, max_row=JR0+29)
chart.add_data(data, titles_from_data=True); chart.set_categories(cats)
chart.y_axis.title="LED cumulees"; chart.x_axis.delete=False; chart.y_axis.delete=False
d.add_chart(chart, "A54")

# ordre des onglets : DASHBOARD, PIPELINE, PARAMETRES
order = ["DASHBOARD","PIPELINE","PARAMETRES"]
wb._sheets.sort(key=lambda s: order.index(s.title))
wb.active = 0
wb.save(OUT)
print("OK ->", OUT)
