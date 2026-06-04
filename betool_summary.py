# -*- coding: utf-8 -*-
"""
betool_summary.py — Cockpit LED 30k
Lit l'export BETOOL (Excel) et ajoute les stats pipeline dans public_data.json.

Usage :
    python betool_summary.py data/betool.xlsx
    python betool_summary.py data/betool.xlsx --output public_data.json
"""
from __future__ import annotations
import json, os, re, sys
from collections import defaultdict
from datetime import date, datetime

try:
    import openpyxl
except ImportError:
    print("[ERREUR] openpyxl manquant. Lancer : pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ---------------------------------------------------------------------------
# Configuration du pipeline (ordre = ordre de la chaîne)
# ---------------------------------------------------------------------------
PIPELINE_STAGES = [
    {
        "id":          "en_cours",
        "label":       "Installation en cours",
        "short":       "En cours",
        "statuts":     ["Installation en cours", "installation en cours EXPRESS",
                        "Installation en Interne"],
        "urgence":     "normal",
        "conseil":     "Attendre fin d'installation avant action.",
    },
    {
        "id":          "attente_audit",
        "label":       "Attente doc & audit",
        "short":       "Audit ⚡",
        "statuts":     ["Installation Fini - En attente doc et audit",
                        "Installation fini en Interne"],
        "urgence":     "action",
        "conseil":     "Contrôler audits → envoyer au client → commenter auditeurs.",
    },
    {
        "id":          "attente_signature",
        "label":       "Attente signature",
        "short":       "Signature ⏳",
        "statuts":     ["Installation fini - En attente signature doc"],
        "urgence":     "relance",
        "conseil":     "Rappeler le client pour signature du dossier.",
    },
    {
        "id":          "modif_audit",
        "label":       "Modif audit avant dépôt",
        "short":       "Modif 🔴",
        "statuts":     ["Installation Fini - Doc ok - Modif Audit"],
        "urgence":     "urgent",
        "conseil":     "Corriger les audits → envoyer à Total Energies.",
    },
    {
        "id":          "depose",
        "label":       "Déposé ✅",
        "short":       "Déposé",
        "statuts":     ["Déposé"],
        "urgence":     "ok",
        "conseil":     "En attente validation + paiement Total Energies.",
    },
]

# Index de recherche rapide statut → étape
_STATUT_TO_STAGE = {}
for _s in PIPELINE_STAGES:
    for _st in _s["statuts"]:
        _STATUT_TO_STAGE[_st] = _s["id"]


# ---------------------------------------------------------------------------
# Lecture Excel
# ---------------------------------------------------------------------------

def load_betool(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) if h is not None else "" for h in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:]]


def inspect_betool(path: str):
    """Affiche les colonnes + une ligne d'exemple pour vérifier le mapping."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h) if h is not None else "" for h in rows[0]]
    print(f"\n{'='*60}\nINSPECTION : {path}\n{'='*60}")
    print(f"{len(rows)-1} lignes · {len(headers)} colonnes\n")
    print("COLONNES détectées :")
    for i, h in enumerate(headers, 1):
        print(f"  {i:>3}. {h}")
    # Vérifier les colonnes clés
    print("\nCOLONNES CLÉS (utilisées par le script) :")
    checks = {
        "Statut (→ étape)":        ["Statut"],
        "LED (→ nb LED)":          ["Nombre de points lumineux ?"],
        "N° dossier (→ ref)":      ["Numéro de dossier", "Numero de dossier", "N° dossier",
                                    "Référence", "Reference", "Ref", "ID", "Id", "Code dossier",
                                    "Numéro de commande Waresito", "Numero de commande Waresito"],
        "Date MAJ (→ âge)":        ["Last updateTime"],
    }
    for label, candidates in checks.items():
        found = next((c for c in candidates if c in headers), None)
        mark = f"✅ '{found}'" if found else "❌ MANQUANTE — ajouter le nom exact dans betool_summary.py"
        print(f"  {label:<22} {mark}")
    if rows and len(rows) > 1:
        print("\nEXEMPLE (1re ligne de données) :")
        for h, v in list(zip(headers, rows[1]))[:12]:
            print(f"  {h:<35} = {v}")


# ---------------------------------------------------------------------------
# Calcul pipeline
# ---------------------------------------------------------------------------

def compute_pipeline(records: list[dict]) -> dict:
    today = date.today()

    # Accumulateurs par stage_id — n, led, ages, dossiers (anon)
    acc: dict[str, dict] = {
        s["id"]: {"n": 0, "led": 0, "ages": [], "dossiers": []}
        for s in PIPELINE_STAGES
    }

    for rec in records:
        statut = rec.get("Statut") or ""
        stage_id = _STATUT_TO_STAGE.get(statut)
        if stage_id is None:
            continue

        # LED
        raw_led = rec.get("Nombre de points lumineux ?")
        try:
            led = int(float(str(raw_led))) if raw_led not in (None, "") else 0
        except (ValueError, TypeError):
            led = 0

        # Référence de recherche équipe — non-PII (jamais nom client / SIRET).
        # L'export BETOOL n'a pas de "numéro de dossier" : on utilise le
        # numéro de commande Waresito (entier, cherchable dans l'outil).
        # SÉCURITÉ : certaines cellules Waresito contiennent du texte libre
        # (initiales, notes type "récup marchandise"). On n'extrait QUE le
        # premier nombre propre ; tout texte est jeté pour éviter toute fuite.
        ref = ""
        for key in ("Numéro de dossier", "Numero de dossier", "N° dossier",
                    "Référence", "Reference", "Ref", "ID", "Id", "Code dossier",
                    "Numéro de commande Waresito", "Numero de commande Waresito"):
            if rec.get(key) not in (None, ""):
                raw = rec[key]
                if isinstance(raw, float) and raw.is_integer():
                    ref = str(int(raw))
                else:
                    m = re.search(r"\d{3,}", str(raw))  # premier nombre ≥3 chiffres
                    ref = m.group(0) if m else ""
                break

        # Âge depuis dernière mise à jour
        upd = rec.get("Last updateTime")
        age = None
        if upd:
            try:
                d = upd.date() if hasattr(upd, "date") else datetime.strptime(str(upd)[:10], "%Y-%m-%d").date()
                age = (today - d).days
            except Exception:
                pass

        acc[stage_id]["n"]   += 1
        acc[stage_id]["led"] += led
        if age is not None:
            acc[stage_id]["ages"].append(age)
        acc[stage_id]["dossiers"].append({"ref": ref, "led": led, "age_days": age})

    # Construction de la réponse
    etapes = []
    total_actif_n   = 0
    total_actif_led = 0

    for s in PIPELINE_STAGES:
        a   = acc[s["id"]]
        avg_age = round(sum(a["ages"]) / len(a["ages"])) if a["ages"] else None
        etape = {
            "id":      s["id"],
            "label":   s["label"],
            "short":   s["short"],
            "urgence": s["urgence"],
            "conseil": s["conseil"],
            "n":       a["n"],
            "led":     a["led"],
            "age_moy": avg_age,
        }
        etapes.append(etape)
        if s["id"] != "depose":
            total_actif_n   += a["n"]
            total_actif_led += a["led"]

    # Zone d'action immédiate (hors "en cours" et "déposé")
    action_ids  = {"attente_audit", "attente_signature", "modif_audit"}
    action_n    = sum(acc[i]["n"]   for i in action_ids)
    action_led  = sum(acc[i]["led"] for i in action_ids)

    depose      = acc["depose"]
    total_led   = depose["led"] + total_actif_led
    pct_depose  = round(depose["led"] / total_led * 100) if total_led else 0

    # Taux de pose : dossiers "posés" (hors en_cours) sur total actif + déposés
    posed_n     = sum(acc[i]["n"] for i in action_ids)   # attente_audit + attente_signature + modif_audit
    depose_n    = depose["n"]
    denom_pose  = total_actif_n + depose_n
    taux_pose_pct = round((posed_n + depose_n) / denom_pose * 100) if denom_pose else 0

    # Dossiers anonymes par stage actionnable (triés par LED desc)
    ACTIONABLE = ["modif_audit", "attente_signature", "attente_audit"]
    dossiers_pipeline = {
        sid: sorted(acc[sid]["dossiers"], key=lambda x: x["led"], reverse=True)
        for sid in ACTIONABLE
    }

    return {
        "generated":        today.strftime("%Y-%m-%d"),
        "total_actif":      total_actif_n,
        "led_actif":        total_actif_led,
        "action_n":         action_n,
        "action_led":       action_led,
        "pct_depose":       pct_depose,
        "taux_pose_pct":    taux_pose_pct,
        "etapes":           etapes,
        "dossiers_pipeline": dossiers_pipeline,
    }


# ---------------------------------------------------------------------------
# Quickwins & snapshot
# ---------------------------------------------------------------------------

_QUICKWIN_META = {
    "modif_audit": {
        "urgence": "urgent", "effort": "< 1h par dossier",
        "action": "Corriger les audits et envoyer à Total Energies immédiatement.",
    },
    "attente_signature": {
        "urgence": "relance", "effort": "Appel/email par dossier",
        "action": "Rappeler les clients en commençant par les dossiers les plus anciens.",
    },
    "attente_audit": {
        "urgence": "action", "effort": "~30 min par audit",
        "action": "Traiter en priorité les plus gros dossiers (plus de LED = plus d'impact).",
    },
}

def compute_quickwins(pipeline: dict) -> list[dict]:
    """Génère la liste des quickwins — utilise les totaux réels des étapes pour n/led."""
    dp      = pipeline.get("dossiers_pipeline", {})
    etapes  = {e["id"]: e for e in pipeline.get("etapes", [])}
    order   = {"urgent": 0, "relance": 1, "action": 2}
    qw      = []
    for stage_id, meta in _QUICKWIN_META.items():
        etape = etapes.get(stage_id, {})
        n_real   = etape.get("n", 0)
        led_real = etape.get("led", 0)
        if n_real == 0:
            continue
        dossiers = dp.get(stage_id, [])
        ages     = [d["age_days"] for d in dossiers if d.get("age_days") is not None]
        old_count = sum(1 for a in ages if a > 7)
        top5_led  = sum(d["led"] for d in dossiers[:5]) if dossiers else 0
        qw.append({
            "rank":        order[meta["urgence"]],
            "stage_id":    stage_id,
            "urgence":     meta["urgence"],
            "n":           n_real,
            "led":         led_real,
            "effort":      meta["effort"],
            "action":      meta["action"],
            "blocage_old": old_count,
            "top5_led":    top5_led,
            "top":         dossiers[:10],
        })
    qw.sort(key=lambda x: (x["rank"], -x["led"]))
    for i, q in enumerate(qw, 1):
        q["rank"] = i
    return qw


def save_snapshot(data: dict, output_path: str = "public_data.json"):
    """Sauvegarde un snapshot daté pour le calcul du delta J-1."""
    today     = date.today().strftime("%Y-%m-%d")
    snap_dir  = os.path.join(os.path.dirname(os.path.abspath(output_path)), "snapshots")
    os.makedirs(snap_dir, exist_ok=True)
    snap_path = os.path.join(snap_dir, f"{today}.json")
    snap = {
        "date":     today,
        "pipeline": data.get("pipeline"),
        "quickwins": data.get("quickwins"),
    }
    with open(snap_path, "w", encoding="utf-8") as f:
        json.dump(snap, f, ensure_ascii=False, indent=2)
    return snap_path


# ---------------------------------------------------------------------------
# Mise à jour public_data.json
# ---------------------------------------------------------------------------

def update_public_json(pipeline: dict, output_path: str = "public_data.json"):
    data = {}
    if os.path.isfile(output_path):
        with open(output_path, encoding="utf-8") as f:
            try:
                data = json.load(f)
            except json.JSONDecodeError:
                pass

    data["pipeline"]          = pipeline
    data["taux_pose_pct"]     = pipeline["taux_pose_pct"]
    data["generated"]         = date.today().strftime("%Y-%m-%d")
    data["dossiers_pipeline"] = pipeline.pop("dossiers_pipeline", {})
    data["quickwins"]         = compute_quickwins(
        {**pipeline, "dossiers_pipeline": data["dossiers_pipeline"]}
    )
    # Provenance : si les dossiers ont de vrais n° → source réelle BETOOL
    has_refs = any(
        d.get("ref") for lst in data["dossiers_pipeline"].values() for d in lst
    )
    data["dossiers_source"] = "betool" if has_refs else "demo"
    snap_path = save_snapshot(data, output_path)

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"[OK] {output_path} mis à jour avec pipeline BETOOL + quickwins")
    print(f"[OK] Snapshot sauvegardé : {snap_path}")
    return data


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    output = "public_data.json"
    for i, a in enumerate(sys.argv[1:]):
        if a == "--output" and i + 2 < len(sys.argv):
            output = sys.argv[i + 2]

    if not args:
        print("Usage: python betool_summary.py <fichier.xlsx> [--output public_data.json]",
              file=sys.stderr)
        print("       python betool_summary.py --inspect <fichier.xlsx>  (vérifier les colonnes)",
              file=sys.stderr)
        sys.exit(1)

    xlsx_path = args[0]
    if not os.path.isfile(xlsx_path):
        print(f"[ERREUR] Fichier introuvable : {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    # Mode inspection : afficher les colonnes et sortir
    if "--inspect" in sys.argv:
        inspect_betool(xlsx_path)
        return

    print(f"Lecture BETOOL : {xlsx_path}")
    records  = load_betool(xlsx_path)
    print(f"  {len(records)} lignes chargées")

    pipeline = compute_pipeline(records)

    print(f"\nPipeline BETOOL :")
    for e in pipeline["etapes"]:
        print(f"  {e['short']:<22} {e['n']:>4} dossiers  {e['led']:>6} LED")
    print(f"\n  Action immédiate : {pipeline['action_n']} dossiers / {pipeline['action_led']} LED")
    print(f"  Déposé : {pipeline['etapes'][-1]['n']} dossiers / {pipeline['etapes'][-1]['led']} LED "
          f"({pipeline['pct_depose']}% du total)")

    update_public_json(pipeline, output)


if __name__ == "__main__":
    main()
