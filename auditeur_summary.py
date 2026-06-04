# -*- coding: utf-8 -*-
"""
auditeur_summary.py — Cockpit LED 30k
Lit l'export BETOOL « PRIME EVOLUTION » (board auditeur, ~2 760 lignes)
et ajoute un bloc `audit_pipeline` dans public_data.json.

Vue complémentaire au CRM : suit la qualité des audits avant dépôt CEE.
Réf dossier = ticket LE-XXXX (non-PII). Ne stocke JAMAIS de PII.

Statuts attendus dans la colonne "Status LED" :
  Étude prête          → prêt à valider/corriger (~2 400)
  Modification à faire → vrais blocages audit (~190) — quickwin prioritaire
  Étude en cours       → audit en traitement
  Étude à réaliser     → à démarrer
  Annulation           → annulé

Colonnes utiles de l'export (noms exacts BETOOL, à adapter si l'export change) :
  Clé ticket       → réf LE-XXXX
  Status LED       → statut audit
  Jetons           → nb LED (ou "Nb LED" selon export)
  Last updateTime  → date dernière modif

Usage :
    python auditeur_summary.py data/prime_evolution.xlsx
    python auditeur_summary.py data/prime_evolution.xlsx --output public_data.json
    python auditeur_summary.py data/prime_evolution.xlsx --inspect
"""
from __future__ import annotations
import json, os, re, sys
from collections import Counter
from datetime import date, datetime

try:
    import openpyxl
except ImportError:
    print("[ERREUR] openpyxl manquant. Lancer : pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# ---------------------------------------------------------------------------
# Pipeline audit (ordre = gravité décroissante / priorité d'action)
# ---------------------------------------------------------------------------
AUDIT_STAGES = [
    {
        "id":      "modif_a_faire",
        "label":   "Modification à faire",
        "short":   "Modif 🔴",
        "statuts": ["Modification à faire", "modification à faire"],
        "urgence": "urgent",
        "conseil": "Corriger les études bloquées → libère les dépôts en attente.",
    },
    {
        "id":      "etude_prête",
        "label":   "Étude prête",
        "short":   "Prêt ✅",
        "statuts": ["Étude prête", "Etude prête", "etude prête", "etude prete", "Etude prete"],
        "urgence": "action",
        "conseil": "Valider et corriger les études prêtes avant envoi au délégataire.",
    },
    {
        "id":      "etude_en_cours",
        "label":   "Étude en cours",
        "short":   "En cours",
        "statuts": ["Étude en cours", "Etude en cours"],
        "urgence": "normal",
        "conseil": "Études en traitement — suivre l'avancement.",
    },
    {
        "id":      "etude_a_realiser",
        "label":   "Étude à réaliser",
        "short":   "À faire",
        "statuts": ["Étude à réaliser", "Etude à réaliser", "étude à réaliser"],
        "urgence": "normal",
        "conseil": "Planifier la réalisation des études restantes.",
    },
    {
        "id":      "annulation",
        "label":   "Annulation",
        "short":   "Annulé",
        "statuts": ["Annulation", "annulation", "Annulé", "annulé"],
        "urgence": "ok",
        "conseil": "",
    },
]

_STATUT_INDEX: dict[str, dict] = {}
for _s in AUDIT_STAGES:
    for _k in _s["statuts"]:
        _STATUT_INDEX[_k.lower().strip()] = _s

# Candidats colonnes (BETOOL peut varier l'orthographe)
_REF_COLS   = ("Clé ticket", "Cle ticket", "Référence", "Reference", "Ticket", "ID")
# NB : dans l'export auditeur, "Jetons" = emoji (♾️/🪙), PAS un nombre de LED.
# Le nb de LED n'existe pas dans cette source → on agrège les dossiers + cellules.
_LED_COLS   = ("Nb LED", "Nombre LED", "Quantite LED", "Quantité LED")
_CELL_COLS  = ("Cellules", "Nombre de cellules", "Nb cellules")
_STATUT_COL = ("Status LED", "Statut LED", "Status", "Statut")
_DATE_COLS  = ("Last updateTime", "Dernière mise à jour", "Date mise à jour",
               "lastUpdateTime", "UpdateTime")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _first(keys, mapping):
    for k in keys:
        if k in mapping:
            return k
    return None


def _num(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r"[^\d,.\-]", "", str(v)).replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _ref(v) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    raw = str(v).strip()
    # Garder LE-XXXX ou tout identifiant numérique ≥3 chiffres
    if re.match(r"^LE-\d+", raw, re.IGNORECASE):
        return raw.upper()
    m = re.search(r"\d{3,}", raw)
    return m.group(0) if m else ""


def _parse_date(v):
    if not v:
        return None
    if isinstance(v, datetime):
        return v.date()
    s = str(v).strip()[:10]
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None


def _age(d: date | None) -> int | None:
    if d is None:
        return None
    return (date.today() - d).days


# ---------------------------------------------------------------------------
# Inspect (mode --inspect)
# ---------------------------------------------------------------------------
def inspect_auditeur(path: str):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        print("[ERREUR] Fichier vide.")
        return

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    print(f"Colonnes ({len(headers)}) :")
    for i, h in enumerate(headers):
        sample = next((str(rows[j][i]) for j in range(1, min(4, len(rows)))
                       if i < len(rows[j]) and rows[j][i] not in (None, "")), "—")
        print(f"  [{i:2d}] {h!r:40s}  ex: {sample}")

    status_col = None
    for h in _STATUT_COL:
        if h in headers:
            status_col = headers.index(h)
            break
    if status_col is not None:
        ctr: Counter = Counter()
        for r in rows[1:]:
            v = r[status_col] if status_col < len(r) else None
            if v:
                ctr[str(v).strip()] += 1
        print(f"\nValeurs '{headers[status_col]}' :")
        for val, cnt in ctr.most_common():
            print(f"  {cnt:5d}  {val!r}")
    else:
        print("\n[AVERTISSEMENT] Colonne 'Status LED' introuvable — vérifier les en-têtes.")


# ---------------------------------------------------------------------------
# Core compute
# ---------------------------------------------------------------------------
def compute_audit_pipeline(path: str) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        print("[ERREUR] Fichier vide.", file=sys.stderr)
        sys.exit(1)

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    h_idx = {h: i for i, h in enumerate(headers)}

    ref_col    = _first(_REF_COLS, h_idx)
    led_col    = _first(_LED_COLS, h_idx)
    cell_col   = _first(_CELL_COLS, h_idx)
    statut_col = _first(_STATUT_COL, h_idx)
    date_col   = _first(_DATE_COLS, h_idx)

    print(f"  Colonnes détectées → ref={ref_col!r} led={led_col!r} "
          f"cellules={cell_col!r} statut={statut_col!r} date={date_col!r}")

    if statut_col is None:
        print("[AVERTISSEMENT] Colonne 'Status LED' absente — tous les dossiers → 'non_reconnu'.",
              file=sys.stderr)

    # Compteurs par stage. NB : pas de nb LED dans cette source → on suit les
    # dossiers (n) et, à défaut, les cellules (zones) comme proxy de volume.
    by_stage: dict[str, dict] = {s["id"]: {"n": 0, "led": 0, "cellules": 0, "refs": [], "dossiers": []}
                                  for s in AUDIT_STAGES}
    by_stage["non_reconnu"] = {"n": 0, "led": 0, "cellules": 0, "refs": [], "dossiers": []}
    today = date.today()

    for row in rows[1:]:
        def cell(col):
            if col is None:
                return None
            i = h_idx.get(col)
            return row[i] if (i is not None and i < len(row)) else None

        statut_raw = str(cell(statut_col) or "").strip()
        stage = _STATUT_INDEX.get(statut_raw.lower(), None)
        sid   = stage["id"] if stage else "non_reconnu"

        led  = int(_num(cell(led_col)))
        cel  = int(_num(cell(cell_col)))
        ref  = _ref(cell(ref_col))
        dt   = _parse_date(cell(date_col))
        age  = _age(dt)

        by_stage[sid]["n"]        += 1
        by_stage[sid]["led"]      += led
        by_stage[sid]["cellules"] += cel
        if ref:
            by_stage[sid]["refs"].append(ref)
        by_stage[sid]["dossiers"].append({"ref": ref, "led": led, "cellules": cel, "age_days": age})

    # Construire les étapes
    etapes = []
    for s in AUDIT_STAGES:
        b = by_stage[s["id"]]
        etapes.append({
            "id":       s["id"],
            "label":    s["label"],
            "short":    s["short"],
            "urgence":  s["urgence"],
            "conseil":  s["conseil"],
            "n":        b["n"],
            "led":      b["led"],
            "cellules": b["cellules"],
        })

    total = sum(e["n"] for e in etapes)
    led_total = sum(e["led"] for e in etapes)
    cellules_total = sum(e["cellules"] for e in etapes)

    # Quickwins audit (modif + etude_prête uniquement)
    quickwins = []
    prio_ids = ["modif_a_faire", "etude_prête"]
    for rank, sid in enumerate(prio_ids, 1):
        b  = by_stage[sid]
        s  = next(x for x in AUDIT_STAGES if x["id"] == sid)
        if b["n"] == 0:
            continue
        doss_sorted = sorted(b["dossiers"], key=lambda x: -(x["cellules"] or 0))
        top5 = doss_sorted[:5]
        quickwins.append({
            "rank":      rank,
            "stage_id":  sid,
            "urgence":   s["urgence"],
            "n":         b["n"],
            "led":       b["led"],
            "cellules":  b["cellules"],
            "action":    s["conseil"],
            "effort":    "~1h" if sid == "modif_a_faire" else "~demi-journée",
            "top5_refs": [d["ref"] for d in top5 if d["ref"]],
        })

    # Dossiers "modif à faire" pour le drawer (top-50)
    modif_doss = sorted(by_stage["modif_a_faire"]["dossiers"],
                        key=lambda x: -(x["cellules"] or 0))[:50]

    non_reco = by_stage["non_reconnu"]["n"]
    result = {
        "generated":      today.strftime("%Y-%m-%d"),
        "cellules_total": cellules_total,
        "source":         "betool_auditeur",
        "total":         total,
        "led_total":     led_total,
        "etapes":        etapes,
        "quickwins":     quickwins,
        "non_reconnu_n": non_reco,
        "modif_dossiers": modif_doss,
    }
    return result


# ---------------------------------------------------------------------------
# Garde-fou PII
# ---------------------------------------------------------------------------
def _assert_no_pii(block: dict):
    blob = json.dumps(block, ensure_ascii=False)
    for pat, label in [
        (r"\b\d{14}\b",              "SIRET"),
        (r"[\w.+-]+@[\w-]+\.[\w.-]+","email"),
        (r"\b0[1-9](?:[\s.-]?\d{2}){4}\b", "téléphone"),
    ]:
        if re.search(pat, blob):
            raise SystemExit(
                f"[STOP PII] {label} détecté dans l'agrégat — écriture annulée."
            )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    output = "public_data.json"
    inspect = "--inspect" in sys.argv

    if "--output" in sys.argv:
        i = sys.argv.index("--output")
        if i + 1 < len(sys.argv):
            output = sys.argv[i + 1]

    if not args:
        print(
            "Usage: python auditeur_summary.py data/prime_evolution.xlsx "
            "[--output public_data.json] [--inspect]",
            file=sys.stderr,
        )
        sys.exit(1)

    path = args[0]
    if not os.path.isfile(path):
        print(f"[ERREUR] Fichier introuvable : {path}", file=sys.stderr)
        sys.exit(1)

    if inspect:
        print(f"=== Inspection {path} ===")
        inspect_auditeur(path)
        return

    print(f"Lecture PRIME EVOLUTION (auditeur) : {path}")
    audit = compute_audit_pipeline(path)
    _assert_no_pii(audit)

    etapes_str = "  ".join(
        f"{e['short']}:{e['n']}" for e in audit["etapes"] if e["n"]
    )
    vol = (f" · {audit['led_total']:,} LED" if audit['led_total']
           else f" · {audit['cellules_total']:,} cellules")
    print(f"  {audit['total']:,} dossiers{vol}".replace(",", " "))
    print(f"  {etapes_str}")
    if audit["non_reconnu_n"]:
        print(f"  ⚠ {audit['non_reconnu_n']} statuts non reconnus "
              f"→ lancer --inspect pour vérifier les valeurs.")
    if audit["quickwins"]:
        qw = audit["quickwins"][0]
        print(f"  ⚡ Top quickwin : {qw['n']} '{qw['stage_id']}' "
              f"({qw['cellules']} cellules)")

    # Merge dans public_data.json
    data: dict = {}
    if os.path.isfile(output):
        with open(output, encoding="utf-8") as f:
            try:
                data = json.load(f)
            except json.JSONDecodeError:
                pass

    data["audit_pipeline"] = audit
    with open(output, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"[OK] {output} enrichi avec le bloc 'audit_pipeline'")


if __name__ == "__main__":
    main()
